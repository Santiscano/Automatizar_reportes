import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font


archiv_excel = pd.read_excel('inputs/supermarket_sales.xlsx') # archivo a analizar
# print(archiv_excel)

tabla_pivote = archiv_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(8)
# print(tabla_pivote)

tabla_pivote.to_excel('sales_2022.xlsx', startrow=4, sheet_name='Report') # crea el archivo excel con el resumen



# crear graficos
wb = load_workbook('sales_2022.xlsx')
pestania = wb['Report']

# sobre que fila y columna se haran los graficos
min_col = wb.active.min_column
max_col = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row

# print('col', min_col)
# print('col', max_col)
# print('row', min_fila)
# print('row', max_fila)

# graficos
barchar = BarChart()

data = Reference(pestania, min_col=min_col+1, max_col=max_col, min_row=min_fila, max_row=max_fila) # sumo 1 en min_col
categoria = Reference(pestania, min_col=min_col, max_col=min_col, min_row=min_fila+1, max_row=max_fila) # sumo 1 en min_row

barchar.add_data(data, titles_from_data=True)
barchar.add_data(data=categoria)

# agregar pestaña al grafico
pestania.add_chart(barchar, 'B12')
barchar.title = 'Ventas'
barchar.style = 5 #este es el color

pestania['A1'] = 'Reporte'
pestania['A2'] = '2022'


pestania['A1'].font = Font('Arial', bold=True, size=20)
pestania['A2'].font = Font('Arial', bold=True, size=15)

wb.save('sales_2022.xlsx')

